import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from datetime import datetime

DATA_DIR  = Path(__file__).parent / 'data'
OUT_PATH  = Path(__file__).parent / 'Meat_Intelligence_Forecasts.xlsx'

# ── Colors ────────────────────────────────────────────────────────────────────
HDR_DARK   = '1F2937'   # dark slate header
HDR_MID    = '374151'   # section sub-header
ROW_ALT    = 'F9FAFB'   # alternate row tint
RED_BG     = 'FEE2E2'
GREEN_BG   = 'D1FAE5'
RED_FT     = 'B91C1C'
GREEN_FT   = '065F46'
GOLD       = 'B45309'
BLUE       = '1D4ED8'

# ── Cut category mapping ───────────────────────────────────────────────────────
BEEF_CUTS    = {'85CL','50CL','insides','flats','eyes','beef_knuckle'}
PORK_CUTS    = {'pork_trim_72','pork_trim_42','pork_belly_13_17','pork_loin',
                'pork_ham_insides','pork_ham_knuckles','pork_ham_outsides',
                'pork_72_fzn','pork_42_fzn'}
POULTRY_CUTS = {'chix_breast_bs','chix_thigh_bs','chix_thigh_bonein',
                'chix_legs','chix_msc','turk_breast_bs','turk_thigh_bs'}

def cat(key):
    if key in BEEF_CUTS:    return 'Beef'
    if key in PORK_CUTS:    return 'Pork'
    if key in POULTRY_CUTS: return 'Poultry'
    return 'Other'

# ── Style helpers ──────────────────────────────────────────────────────────────
def hdr_fill(hex_color):
    return PatternFill('solid', start_color=hex_color, end_color=hex_color)

def thin_border():
    s = Side(style='thin', color='D1D5DB')
    return Border(left=s, right=s, top=s, bottom=s)

def apply_hdr(cell, text, hex_bg=HDR_DARK, size=10):
    cell.value = text
    cell.font  = Font(bold=True, color='FFFFFF', name='Arial', size=size)
    cell.fill  = hdr_fill(hex_bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border()

def apply_data(cell, value, fmt=None, bold=False, color=None, bg=None, align='right'):
    cell.value = value
    cell.font  = Font(name='Arial', size=9, bold=bold,
                      color=color if color else '111827')
    if bg:
        cell.fill = hdr_fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = thin_border()
    if fmt:
        cell.number_format = fmt

FMT_CWT = '"$"#,##0.00"/cwt"'
FMT_LB  = '"$"#,##0.0000"/lb"'
FMT_PCT = '0.00"%"'
FMT_NUM = '#,##0'

# ── Load data ──────────────────────────────────────────────────────────────────
raw       = json.loads((DATA_DIR / 'forecasts.json').read_text(encoding='utf-8'))
forecasts = raw['forecasts']
gen_at    = raw.get('generated_at', '')
rep_date  = raw.get('report_date', '')

HORIZONS = ['1w', '4w', '13w', '26w']

# ── Build flat rows ────────────────────────────────────────────────────────────
def build_rows(keys):
    rows = []
    for k in keys:
        if k not in forecasts:
            continue
        f  = forecasts[k]
        hz = f.get('horizons', {})
        bt = f.get('backtest', {})
        rows.append({
            'cut_key':       k,
            'name':          f.get('name', k),
            'category':      cat(k),
            'source':        f.get('source_report',''),
            'model':         f.get('model',''),
            'history_pts':   f.get('history_points', 0),
            'anchor_cwt':    f.get('anchor_cwt'),
            'anchor_lb':     f.get('anchor_lb'),
            'prior_cwt':     f.get('prior_day_cwt'),
            'delta_cwt':     f.get('delta_day_cwt'),
            'delta_pct':     f.get('delta_day_pct'),
            'lo52_cwt':      f.get('range_52w_low_cwt'),
            'hi52_cwt':      f.get('range_52w_high_cwt'),
            # horizons
            '1w_p10':  hz.get('1w',{}).get('p10_cwt'),
            '1w_p50':  hz.get('1w',{}).get('p50_cwt'),
            '1w_p90':  hz.get('1w',{}).get('p90_cwt'),
            '4w_p10':  hz.get('4w',{}).get('p10_cwt'),
            '4w_p50':  hz.get('4w',{}).get('p50_cwt'),
            '4w_p90':  hz.get('4w',{}).get('p90_cwt'),
            '13w_p10': hz.get('13w',{}).get('p10_cwt'),
            '13w_p50': hz.get('13w',{}).get('p50_cwt'),
            '13w_p90': hz.get('13w',{}).get('p90_cwt'),
            '26w_p10': hz.get('26w',{}).get('p10_cwt'),
            '26w_p50': hz.get('26w',{}).get('p50_cwt'),
            '26w_p90': hz.get('26w',{}).get('p90_cwt'),
        })
    return rows

def write_sheet(ws, rows, title):
    # ── Title bar ──────────────────────────────────────────────────────────────
    ws.merge_cells('A1:AB1')
    c = ws['A1']
    c.value     = f'Jack Link\'s Meat Intelligence — {title}  |  Report Date: {rep_date}  |  Generated: {gen_at}'
    c.font      = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    c.fill      = hdr_fill('111827')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    # ── Column headers (rows 2-3 merged for horizon groups) ───────────────────
    # Row 2: group labels
    # Row 3: individual sub-headers

    COLS = [
        ('A', 'Cut Key',        8),
        ('B', 'Cut Name',       26),
        ('C', 'Category',       8),
        ('D', 'Source Report',  10),
        ('E', 'Model',          14),
        ('F', 'History\nPoints',8),
        ('G', 'Today\n$/cwt',   10),
        ('H', 'Today\n$/lb',    9),
        ('I', 'Prior Day\n$/cwt',10),
        ('J', 'Δ Day\n$/cwt',   9),
        ('K', 'Δ Day\n%',       7),
        ('L', '52W Low\n$/cwt', 10),
        ('M', '52W High\n$/cwt',10),
    ]
    # Horizon group columns start at N (col 14)
    hz_cols = {
        '1W FORECAST':  ('N','O','P'),
        '4W FORECAST':  ('Q','R','S'),
        '13W FORECAST': ('T','U','V'),
        '26W FORECAST': ('W','X','Y'),
    }

    # Fixed col headers rows 2+3 (merged vertically)
    for col_letter, label, width in COLS:
        ws.merge_cells(f'{col_letter}2:{col_letter}3')
        apply_hdr(ws[f'{col_letter}2'], label, HDR_DARK)
        ws.column_dimensions[col_letter].width = width

    # Horizon group headers (row 2 merged across 3 cols, row 3 sub-headers)
    hz_bg_colors = ['1E40AF','065F46','92400E','4C1D95']
    sub_labels   = ['Best (P10)','Base (P50)','Worst (P90)']
    for idx, (grp, (c1, c2, c3)) in enumerate(hz_cols.items()):
        bg = hz_bg_colors[idx]
        ws.merge_cells(f'{c1}2:{c3}2')
        apply_hdr(ws[f'{c1}2'], grp, bg, size=9)
        for cl, lbl in zip((c1, c2, c3), sub_labels):
            apply_hdr(ws[f'{cl}3'], lbl, HDR_MID, size=8)
            ws.column_dimensions[cl].width = 10

    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    # ── Data rows ──────────────────────────────────────────────────────────────
    for ri, row in enumerate(rows):
        r   = ri + 4
        alt = ROW_ALT if ri % 2 == 0 else 'FFFFFF'

        # Delta color logic: cost DOWN = good (green), cost UP = bad (red)
        d   = row['delta_cwt'] or 0
        dbg = GREEN_BG if d <= 0 else RED_BG
        dft = GREEN_FT if d <= 0 else RED_FT

        model_short = 'GBM ✓' if 'gbm' in (row['model'] or '') else 'Parametric'

        cells_data = [
            ('A', row['cut_key'],    None,    False, '374151', alt,  'left'),
            ('B', row['name'],       None,    True,  '111827', alt,  'left'),
            ('C', row['category'],   None,    False, GOLD,     alt,  'center'),
            ('D', row['source'],     None,    False, '6B7280', alt,  'center'),
            ('E', model_short,       None,    False, BLUE,     alt,  'center'),
            ('F', row['history_pts'],FMT_NUM, False, '374151', alt,  'center'),
            ('G', row['anchor_cwt'], FMT_CWT, True,  '111827', alt,  'right'),
            ('H', row['anchor_lb'],  FMT_LB,  False, '374151', alt,  'right'),
            ('I', row['prior_cwt'],  FMT_CWT, False, '6B7280', alt,  'right'),
            ('J', row['delta_cwt'],  '"$"#,##0.00;-"$"#,##0.00', True, dft, dbg, 'right'),
            ('K', (row['delta_pct'] or 0)/100 if row['delta_pct'] is not None else None,
                  '0.00%;-0.00%', True, dft, dbg, 'right'),
            ('L', row['lo52_cwt'],   FMT_CWT, False, '374151', alt,  'right'),
            ('M', row['hi52_cwt'],   FMT_CWT, False, '374151', alt,  'right'),
            # 1w
            ('N', row['1w_p10'],     FMT_CWT, False, GREEN_FT, GREEN_BG, 'right'),
            ('O', row['1w_p50'],     FMT_CWT, True,  '111827', alt,  'right'),
            ('P', row['1w_p90'],     FMT_CWT, False, RED_FT,   RED_BG,   'right'),
            # 4w
            ('Q', row['4w_p10'],     FMT_CWT, False, GREEN_FT, GREEN_BG, 'right'),
            ('R', row['4w_p50'],     FMT_CWT, True,  '111827', alt,  'right'),
            ('S', row['4w_p90'],     FMT_CWT, False, RED_FT,   RED_BG,   'right'),
            # 13w
            ('T', row['13w_p10'],    FMT_CWT, False, GREEN_FT, GREEN_BG, 'right'),
            ('U', row['13w_p50'],    FMT_CWT, True,  '111827', alt,  'right'),
            ('V', row['13w_p90'],    FMT_CWT, False, RED_FT,   RED_BG,   'right'),
            # 26w
            ('W', row['26w_p10'],    FMT_CWT, False, GREEN_FT, GREEN_BG, 'right'),
            ('X', row['26w_p50'],    FMT_CWT, True,  '111827', alt,  'right'),
            ('Y', row['26w_p90'],    FMT_CWT, False, RED_FT,   RED_BG,   'right'),
        ]

        for col, val, fmt, bold, color, bg, align in cells_data:
            apply_data(ws[f'{col}{r}'], val, fmt, bold, color, bg, align)

        ws.row_dimensions[r].height = 16

    # ── Legend row ─────────────────────────────────────────────────────────────
    leg_row = len(rows) + 5
    ws.merge_cells(f'A{leg_row}:Y{leg_row}')
    lc = ws[f'A{leg_row}']
    lc.value = (
        'Legend:  Best (P10) = Optimistic scenario (cost likely above this 10% of the time)  |  '
        'Base (P50) = Most likely price  |  '
        'Worst (P90) = Pessimistic scenario (cost likely below this 90% of the time)  |  '
        'Green Δ = Price declined (favorable for procurement)  |  '
        'Red Δ = Price increased (cost pressure)  |  '
        'GBM ✓ = Machine-learning model beat seasonal baseline  |  '
        'Parametric = Statistical fallback model'
    )
    lc.font      = Font(name='Arial', size=8, italic=True, color='6B7280')
    lc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[leg_row].height = 28

    # ── Freeze panes ───────────────────────────────────────────────────────────
    ws.freeze_panes = 'B4'

# ── Build workbook ─────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)   # remove default blank sheet

all_keys     = list(forecasts.keys())
beef_keys    = [k for k in all_keys if k in BEEF_CUTS]
pork_keys    = [k for k in all_keys if k in PORK_CUTS]
poultry_keys = [k for k in all_keys if k in POULTRY_CUTS]
other_keys   = [k for k in all_keys if k not in BEEF_CUTS | PORK_CUTS | POULTRY_CUTS]

CUT_FORECAST_KEYS = ['insides', 'flats', '50CL', '85CL',
                     'pork_ham_insides', 'pork_ham_knuckles', 'pork_ham_outsides']

sheets = [
    ('Cut Forecast',  CUT_FORECAST_KEYS),
    ('All Cuts',      all_keys),
    ('Beef Cuts',     beef_keys),
    ('Pork Cuts',     pork_keys),
    ('Poultry Cuts',  poultry_keys),
]
if other_keys:
    sheets.append(('Other Cuts', other_keys))

for sheet_title, keys in sheets:
    ws = wb.create_sheet(sheet_title)
    write_sheet(ws, build_rows(keys), sheet_title)

# ── Notes sheet ────────────────────────────────────────────────────────────────
ns = wb.create_sheet('Notes & Sources')
notes = [
    ('Field',              'Description'),
    ('Today $/cwt',        'USDA AMS weighted average price per hundredweight (100 lbs) — pulled from live PDF report'),
    ('Today $/lb',         'Today $/cwt divided by 100'),
    ('Prior Day $/cwt',    'Previous trading day closing price from 5-year history database'),
    ('Δ Day $/cwt',        'Today minus prior day price. Green = favorable (price fell), Red = cost pressure (price rose)'),
    ('Δ Day %',            'Percentage change from prior day'),
    ('52W Low / High',     'Min and max price over trailing 365 calendar days'),
    ('Best (P10)',         'Optimistic forecast — model predicts only 10% chance price falls below this'),
    ('Base (P50)',         'Base case — median forecast, most likely outcome'),
    ('Worst (P90)',        'Pessimistic forecast — model predicts 90% chance price falls below this'),
    ('Model: GBM ✓',      'Quantile Gradient Boosting model — trained on 2023+ history, passed walk-forward backtest vs seasonal baseline'),
    ('Model: Parametric',  'Statistical seasonal-drift fallback — used when history is thin or GBM did not beat baseline'),
    ('History Points',     'Number of daily price observations used to train the forecast model'),
    ('',                   ''),
    ('Data Sources',       ''),
    ('LM_XB401',           'USDA AMS National Daily Boneless Processing Beef & Beef Trimmings — ams.usda.gov/mnreports/ams_2451.pdf'),
    ('LM_XB403',           'USDA AMS National Daily Boxed Beef Cutout & Cuts — ams.usda.gov/mnreports/ams_2453.pdf'),
    ('LM_XB405',           'USDA AMS National 5-Day Rolling Cutter Cow Cutout & 100VL — ams.usda.gov/mnreports/ams_2455.pdf'),
    ('LM_PK602',           'USDA AMS National Daily Pork FOB Plant — ams.usda.gov/mnreports/ams_2498.pdf'),
    ('LM_PK610',           'USDA AMS National Weekly Pork Summary (includes frozen trim)'),
    ('AMS_3646',           'USDA AMS National Weekly Chicken Parts Report'),
    ('AMS_3647',           'USDA AMS National Weekly Turkey Report'),
    ('',                   ''),
    ('Report Date:',       rep_date),
    ('Generated At:',      gen_at),
    ('Prepared For:',      "Jack Link's Procurement & Supply Chain"),
]
apply_hdr(ns['A1'], 'Field', HDR_DARK)
apply_hdr(ns['B1'], 'Description', HDR_DARK)
ns.column_dimensions['A'].width = 22
ns.column_dimensions['B'].width = 80
for ri, (f, d) in enumerate(notes[1:], start=2):
    bold = (f in ('Data Sources','Report Date:','Generated At:','Prepared For:'))
    apply_data(ns[f'A{ri}'], f, bold=bold, align='left', bg='F9FAFB' if ri%2==0 else 'FFFFFF')
    apply_data(ns[f'B{ri}'], d, align='left',             bg='F9FAFB' if ri%2==0 else 'FFFFFF')
    ns.row_dimensions[ri].height = 15

wb.save(OUT_PATH)
print(f'Saved: {OUT_PATH}')
